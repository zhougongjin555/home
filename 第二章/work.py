import os
from xml.etree import ElementTree as et
from openpyxl import load_workbook
import requests
import configparser
#
#
# base_dir = os.path.dirname(__file__)
# file_path = os.path.join(base_dir, 'login.csv')
#
# # 注册
# with open(file_path, mode='w', encoding='utf-8') as f:
#     while True:
#         msg = input("请注册你的用户名和密码(A-B)(输入Q停止)：")
#         msg = msg.strip('') # 删除空格
#         if msg == 'Q' or msg == 'q':
#             break
#         f.write(msg)
#
# # 登陆验证
# with open(file_path, mode='r', encoding='utf-8') as f:
#     mes = input("请输入你的用户名以及密码(A-B)：")
#     for line in f:
#         if mes == line:
#             print('登陆成功')
#         else:
#             exit(0)
#
#



# while True:
#     city = input("请输入城市（Q/q退出）：")
#     if city.upper() == "Q":
#         break
#     url = "http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={}".format(city)
#     res = requests.get(url=url)
#     # print(res.text)
#
# root = et.XML(res.text) # 获取根节点
# country = root.find('string')  # 找节点
# print(country.tag, country.attrib, country.text)# 找节点里面的标签，值，以及文本信息
#
#
# # 以下操作都要先操作再保存才能实现修改
# country.set('a', 'value')# 修改节点的值
# country.text = 'str'  # 修改节点的文本信息
#
# root.remove(country) # 删除节点
#
# #保存
# root.write('a.xml', encoding='utf-8')


base_dir = os.path.dirname(__file__)
file_dir = os.path.join(base_dir, 'config.ini')
excel_dir = os.path.join(base_dir, 'config.xlsx')

config = configparser.ConfigParser()
config.read(file_dir, encoding='utf-8')

print(config.sections()) # 获取所有节点
print(config.items('mysqld')) # 获取节点的键值对
print(config.get('mysqld', 'datadir')) # 获取特定节点的特定键下面的值


wb = load_workbook(excel_dir)
sheet = wb.worksheets[0]
cell1, cell2, cell3 = sheet.cell(1,1), sheet.cell(1,2), sheet.cell(1,3)
cell1.value, cell2.value, cell3.value = 'root', 'key', 'value'
# wb.save(excel_dir)
i = 2
for root in config.sections():
    for key,value in config.items(root):
        cell1, cell2, cell3 = sheet.cell(i, 1), sheet.cell(i, 2), sheet.cell(i, 3)
        cell1.value, cell2.value, cell3.value = root, key, value
        i += 1
wb.save(excel_dir)